from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
book = load_workbook('jobs/qual - Copy - Copy.xlsx')
array_names = ["Machine Learning and AI", 'Hardware', 'Software and Services', 'Design', 'Operations and Supply Chain', 'Marketing', 'Corporate Functions', 'Apple Retail', 'Sales and Business Development', 'Support and Service', 'Students']

sheet=book.active
def make_excel_coor(row,column):
    if column==1:
        return "A"+str(row)+":"+"A"+str(row)
    else:
        return "B"+str(row)+":"+"B"+str(row)
for j in range(len(array_names)):
    sheet1 = book[array_names[j]]
    sheet1.cell(1,1).font = Font(size=24, bold=True)
    sheet1.cell(1,2).font = Font(size=24, bold=True)
    column1=set()
    column2=set()
    seti=list()
    sheet1.column_dimensions['A'].width=8.11
    sheet1.column_dimensions['B'].width=8.11
    for i in range(2,3000):
        sheet1.row_dimensions[i].height=14.4
        if sheet1.cell(row=i, column=1).value not in column1:
            column1.add(sheet1.cell(row=i, column=1).value)
        else:
            sheet1.move_range(make_excel_coor(i,1),0,3)
            #print("column1, row:",i)

        if sheet1.cell(row=i, column=2).value not in column2:
            column2.add(sheet1.cell(row=i, column=2).value)
        else:
            #print("column2, row:", i)
            sheet1.move_range(make_excel_coor(i,2),0,3) 
        if sheet1.cell(row = i, column = 1).value == None and sheet1.cell(row = i, column = 2).value == None:
            seti.append(i)
    sheet1.delete_cols(4,5)
    f=0
    print(seti)
    for i in range(len(seti)):
        sheet1.delete_rows(seti[i]-f,1)
        f+=1
    print("done")
book.save('jobs/qual - Copy - Copy.xlsx')