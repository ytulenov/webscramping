from selenium import webdriver
from selenium.webdriver.common.by import By

from openpyxl import Workbook, load_workbook
book = load_workbook('jobs/qual.xlsx')
sheet=book.active
sheet.title="Machine Learning and AI"
sheet.append(['Key Qualifications', 'Description'])
sheet_key_row=2
sheet_key_column=1
sheet_desc_row=2
sheet_desc_column=2
book.create_sheet('Hardware')
sheet1 = book['Hardware']
sheet1.append(['Key Qualifications', 'Description'])
sheet1_key_row=2
sheet1_key_column=1
sheet1_desc_row=2
sheet1_desc_column=2
book.create_sheet('Software and Services')
sheet2 = book['Software and Services']
sheet2.append(['Key Qualifications', 'Description'])
sheet2_key_row=2
sheet2_key_column=1
sheet2_desc_row=2
sheet2_desc_column=2

book.create_sheet('Design')
sheet3 = book['Design']
sheet3.append(['Key Qualifications', 'Description'])
sheet3_key_row=2
sheet3_key_column=1
sheet3_desc_row=2
sheet3_desc_column=2

book.create_sheet('Operations and Supply Chain')
sheet4 = book['Operations and Supply Chain']
sheet4.append(['Key Qualifications', 'Description'])
sheet4_key_row=2
sheet4_key_column=1
sheet4_desc_row=2
sheet4_desc_column=2

book.create_sheet('Marketing')
sheet5 = book['Marketing']
sheet5.append(['Key Qualifications', 'Description'])
sheet5_key_row=2
sheet5_key_column=1
sheet5_desc_row=2
sheet5_desc_column=2

book.create_sheet('Corporate Functions')
sheet6 = book['Corporate Functions']
sheet6.append(['Key Qualifications', 'Description'])
sheet6_key_row=2
sheet6_key_column=1
sheet6_desc_row=2
sheet6_desc_column=2

book.create_sheet('Apple Retail')
sheet7 = book['Apple Retail']
sheet7.append(['Key Qualifications', 'Description'])
sheet7_key_row=2
sheet7_key_column=1
sheet7_desc_row=2
sheet7_desc_column=2

book.create_sheet('Sales and Business Development')
sheet8 = book['Sales and Business Development']
sheet8.append(['Key Qualifications', 'Description'])
sheet8_key_row=2
sheet8_key_column=1
sheet8_desc_row=2
sheet8_desc_column=2

book.create_sheet('Support and Service')
sheet9 = book['Support and Service']
sheet9.append(['Key Qualifications', 'Description'])
sheet9_key_row=2
sheet9_key_column=1
sheet9_desc_row=2
sheet9_desc_column=2

book.create_sheet('Students')
sheet10 = book['Students']
sheet10.append(['Key Qualifications', 'Description'])
sheet10_key_row=2
sheet10_key_column=1
sheet10_desc_row=2
sheet10_desc_column=2






driver=webdriver.Chrome()
koko=webdriver.Chrome()
#ertyui=0
checker=0
link='https://jobs.apple.com/en-ca/search?team=business-intelligence-and-analytics-OPMFG-BIA%20business-process-management-OPMFG-BPM%20supply-demand-management-and-npi-readiness-OPMFG-SDMNR%20retail-and-e-commerce-fulfillment-OPMFG-RECF%20logistics-and-supply-chain-OPMFG-SCL%20sales-planning-and-operations-OPMFG-SPO%20procurement-OPMFG-PRC%20manufacturing-and-operations-engineering-OPMFG-MFGE%20quality-engineering-OPMFG-QE%20supplier-responsibility-OPMFG-SR%20program-management-OPMFG-PRMGMT%20information-systems-and-technology-CORSV-IT%20finance-CORSV-FIN%20legal-CORSV-LEG%20people-CORSV-HR%20learning-and-development-CORSV-LRNDV%20global-security-CORSV-GLSEC%20information-security-CORSV-INFOSEC%20environment-and-social-initiatives-CORSV-ENSI%20policy-and-government-affairs-CORSV-GOV%20real-estate-and-development-CORSV-REFAC%20dining-and-food-services-CORSV-DFS%20administration-CORSV-ADMIN%20global-retail-support-CORSV-GRS%20acoustic-technologies-HRDWR-ACT%20analog-and-digital-design-HRDWR-ADD%20architecture-HRDWR-ARCH%20battery-engineering-HRDWR-BE%20camera-technologies-HRDWR-CAM%20display-technologies-HRDWR-DISP%20engineering-project-management-HRDWR-EPM%20environmental-technologies-HRDWR-ENVT%20health-technology-HRDWR-HT%20machine-learning-and-ai-HRDWR-MCHLN%20mechanical-engineering-HRDWR-ME%20process-engineering-HRDWR-PE%20reliability-engineering-HRDWR-REL%20sensor-technologies-HRDWR-SENT%20silicon-technologies-HRDWR-SILT%20system-design-and-test-engineering-HRDWR-SDE%20wireless-hardware-HRDWR-WT%20business-development-SLDEV-BUSDEV%20account-management-SLDEV-CC%20apple-store-sales-SLDEV-ARS%20retail-partner-sales-SLDEV-CRC%20sales-planning-and-operations-SLDEV-SO%20field-and-solutions-engineering-SLDEV-FSE%20apps-and-frameworks-SFTWR-AF%20cloud-and-infrastructure-SFTWR-CLD%20core-operating-systems-SFTWR-COS%20devops-and-site-reliability-SFTWR-DSR%20engineering-project-management-SFTWR-EPM%20information-systems-and-technology-SFTWR-ISTECH%20machine-learning-and-ai-SFTWR-MCHLN%20security-and-privacy-SFTWR-SEC%20software-quality-automation-and-tools-SFTWR-SQAT%20wireless-software-SFTWR-WSFT%20internships-STDNT-INTRN%20corporate-STDNT-CORP%20apple-store-STDNT-ASTR%20apple-store-leader-program-STDNT-ASLP%20apple-retail-partner-store-STDNT-ARPS%20apple-support-college-program-STDNT-ACCP%20apple-campus-leader-STDNT-ACR%20machine-learning-infrastructure-MLAI-MLI%20deep-learning-and-reinforcement-learning-MLAI-DLRL%20natural-language-processing-and-speech-technologies-MLAI-NLP%20computer-vision-MLAI-CV%20applied-research-MLAI-AR%20online-support-CUST-ONSPT%20technical-support-and-customer-support-CUST-ACCS%20apple-store-support-CUST-ACRCC%20applecare-business-development-CUST-SSBD%20service-channel-management-and-operations-CUST-SCMO%20services-marketing-MKTG-SVCM%20product-marketing-MKTG-PM%20marketing-communications-MKTG-MKTCM%20corporate-communications-MKTG-CRPCM%20industrial-design-DESGN-ID%20human-interface-design-DESGN-HID%20communications-design-DESGN-CMD%20sales-APPST-ARSS%20support-APPST-ARSCS%20leadership-APPST-ARSLD'
while checker==0:
    checker=1
    driver.get(link)
    driver.implicitly_wait(5)
    button = driver.find_elements(By.CLASS_NAME, 'table--advanced-search__title')
    for but in button:
        try:
            #print(ertyui+1)
            #print(but.text)
            #print(but.get_attribute("href"))
            #print()
            koko.get(but.get_attribute("href"))
            job_type = koko.find_element(By.ID, 'job-team-name')
            key_qualifications = koko.find_elements(By.ID, 'jd-key-qualifications')
            if job_type.text=="Apple Retail":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet7.cell(row=sheet7_key_row, column=sheet7_key_column).value=k.text
                    sheet7_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet7.cell(row=sheet7_desc_row, column=sheet7_desc_column).value=k.text
                    sheet7_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet7.cell(row=sheet7_key_row, column=sheet7_key_column).value=k.text
                        sheet7_key_row+=1
                        #print()
                except:
                    continue
                #print()    
                #print("|||||||||||||||||||||||||||||")
                #print()
                #print()
                #print()
            if job_type.text=="Students":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet10.cell(row=sheet10_key_row, column=sheet10_key_column).value=k.text
                    sheet10_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet10.cell(row=sheet10_desc_row, column=sheet10_desc_column).value=k.text
                    sheet10_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet10.cell(row=sheet10_key_row, column=sheet10_key_column).value=k.text
                        sheet10_key_row+=1
                        #print()
                except:
                    continue
                #print()    
                #print("|||||||||||||||||||||||||||||")
                #print()
                #print()
                #print()
            if job_type.text=="Support and Service":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet9.cell(row=sheet9_key_row, column=sheet9_key_column).value=k.text
                    sheet9_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet9.cell(row=sheet9_desc_row, column=sheet9_desc_column).value=k.text
                    sheet9_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet9.cell(row=sheet9_key_row, column=sheet9_key_column).value=k.text
                        sheet9_key_row+=1
                        #print()
                except:
                    continue
                #print()    
                #print("|||||||||||||||||||||||||||||")
                #print()
                #print()
                #print()
            if job_type.text=="Machine Learning and AI":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet.cell(row=sheet_key_row, column=sheet_key_column).value=k.text
                    sheet_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet.cell(row=sheet_desc_row, column=sheet_desc_column).value=k.text
                    sheet_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet.cell(row=sheet_key_row, column=sheet_key_column).value=k.text
                        sheet_key_row+=1
                        #print()
                except:
                    continue
                #print()    
                #print("|||||||||||||||||||||||||||||")
                #print()
                #print()
                #print()
            if job_type.text=="Hardware":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet1.cell(row=sheet1_key_row, column=sheet1_key_column).value=k.text
                    sheet1_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet1.cell(row=sheet1_desc_row, column=sheet1_desc_column).value=k.text
                    sheet1_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet1.cell(row=sheet1_key_row, column=sheet1_key_column).value=k.text
                        sheet1_key_row+=1
                        #print()
                except:
                    continue
                #print()    
                #print("|||||||||||||||||||||||||||||")
                #print()
                #print()
                #print()
            if job_type.text=="Software and Services":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet2.cell(row=sheet2_key_row, column=sheet2_key_column).value=k.text
                    sheet2_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet2.cell(row=sheet2_desc_row, column=sheet2_desc_column).value=k.text
                    sheet2_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet2.cell(row=sheet2_key_row, column=sheet2_key_column).value=k.text
                        sheet2_key_row+=1
                        #print()
                except:
                    continue
                #print()    
                #print("|||||||||||||||||||||||||||||")
                #print()
                #print()
                #print()
            if job_type.text=="Design":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet3.cell(row=sheet3_key_row, column=sheet3_key_column).value=k.text
                    sheet3_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet3.cell(row=sheet3_desc_row, column=sheet3_desc_column).value=k.text
                    sheet3_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet3.cell(row=sheet3_key_row, column=sheet3_key_column).value=k.text
                        sheet3_key_row+=1
                        #print()
                except:
                    continue
                #print()    
                #print("|||||||||||||||||||||||||||||")
                #print()
                #print()
                #print() 
            if job_type.text=="Operations and Supply Chain":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet4.cell(row=sheet4_key_row, column=sheet4_key_column).value=k.text
                    sheet4_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet4.cell(row=sheet4_desc_row, column=sheet4_desc_column).value=k.text
                    sheet4_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet4.cell(row=sheet4_key_row, column=sheet4_key_column).value=k.text
                        sheet4_key_row+=1
                        #print()
                except:
                    continue  
            if job_type.text=="Marketing":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet5.cell(row=sheet5_key_row, column=sheet5_key_column).value=k.text
                    sheet5_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet5.cell(row=sheet5_desc_row, column=sheet5_desc_column).value=k.text
                    sheet5_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet5.cell(row=sheet5_key_row, column=sheet5_key_column).value=k.text
                        sheet5_key_row+=1
                        #print()
                except:
                    continue 
            if job_type.text=="Corporate Functions":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet6.cell(row=sheet6_key_row, column=sheet6_key_column).value=k.text
                    sheet6_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet6.cell(row=sheet6_desc_row, column=sheet6_desc_column).value=k.text
                    sheet6_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet6.cell(row=sheet6_key_row, column=sheet6_key_column).value=k.text
                        sheet6_key_row+=1
                        #print()
                except:
                    continue
            if job_type.text=="Sales and Business Development":
                #print("Apple Retail")
                for k in key_qualifications:
                    #print("Key qualifications")
                    sheet8.cell(row=sheet8_key_row, column=sheet8_key_column).value=k.text
                    sheet8_key_row+=1
                    #print()
                description = koko.find_elements(By.ID, 'jd-description')
                for k in description:
                    #print("Description")
                    sheet8.cell(row=sheet8_desc_row, column=sheet8_desc_column).value=k.text
                    sheet8_desc_row+=1
                    #print()
                try:
                    additionalreq = koko.find_elements(By.ID, 'jd-additional-requirements')
                    for k in additionalreq:
                        #print("Additional Requirements")
                        sheet8.cell(row=sheet8_key_row, column=sheet8_key_column).value=k.text
                        sheet8_key_row+=1
                        #print()
                except:
                    continue
        except:
            continue
    next_page = driver.find_elements(By.TAG_NAME, 'a')
    a="Next Page"        
    checker = driver.find_element(By.ID, 'page-number')
    ko = driver.find_elements(By.CLASS_NAME, 'pageNumber')
    for k in ko:
        fqwd=k.text
    fqwd = int(fqwd)
    
    if fqwd == int(checker.get_attribute('value')):
        print("Last Page")
    else:
        for i in next_page:
            try: 
                    f=str(i.get_attribute("href"))
                    #print(str(i.get_attribute("href")))
                    if i.text==a:
                        link=f
                        checker=0
                        #print(f)
                        #print("|||||||||")
                        #print("|||||||||")
                        #print("|||||||||")
                        #print("|||||||||")
                        #print("|||||||||")
                        #print("|||||||||")
                        break
            except:
                pass
    #ertyui+=1

book.save('jobs/qual.xlsx')



